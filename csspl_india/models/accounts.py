from odoo import fields,api,_,models
from odoo.exceptions import ValidationError, UserError
from num2words import num2words
import re
from odoo.exceptions import ValidationError, UserError
from num2words import num2words
import re
import datetime
import io
import base64
from datetime import datetime,time
import xlsxwriter
from io import BytesIO
import pytz
import xlwt
from openpyxl import Workbook
from babel.numbers import format_decimal
from collections import defaultdict


class AccountMoveInherit(models.Model):
    _inherit = 'account.move'

    po_no = fields.Char('PO No')
    po_date = fields.Date('PO Date')
    service_date = fields.Date('Service Date')
    partner_id_line = fields.Many2one('res.partner', string="Partner Line")
    amt = fields.Char(compute='amt_in_words', string='amt')
    bank_name = fields.Char('Bank')
    cheque_no = fields.Char(string="Check No")
    cheque_dt = fields.Date(string="Cheque Date")
    analytic_distribution_id = fields.Many2one('account.analytic.account', 'Analytic Distribution', tracking=True)
    analytic_distribution_journal_id = fields.Many2one('account.analytic.account', 'Analytic Distribution',
                                                       tracking=True)
    analytics_plan_id = fields.Many2one('account.analytic.plan', string='Analytics Plan', tracking=True)

    # For addeding 3 fields in gstr1 tool invoices tree view
    igst = fields.Float(string="IGST", compute="_compute_gst_amount", store=True)
    cgst = fields.Float(string="CGST")
    sgst = fields.Float(string="SGST")
    gst_amount = fields.Float(string="GST")
    tds_amount = fields.Float(string="TDS")
    batch_payment_id = fields.Many2one(related='origin_payment_id.batch_payment_id')

    # Force the journal to show in invoice, even if one option exists
    @api.depends('suitable_journal_ids')
    def _compute_show_journal(self):
        for move in self:
            move.show_journal = len(move.suitable_journal_ids) >= 1

    @api.constrains('line_ids', 'invoice_line_ids')
    def _check_non_zero_entries(self):
        for move in self:
            total_debit = 0
            total_credit = 0
            if move.move_type != 'entry':
                if move.invoice_line_ids:
                    total_debit = sum(move.invoice_line_ids.mapped('price_unit'))
                    total_credit = sum(move.invoice_line_ids.mapped('price_unit'))
            if move.move_type == 'entry':
                if move.line_ids:
                    total_debit = sum(move.line_ids.mapped('debit'))
                    total_credit = sum(move.line_ids.mapped('credit'))
            if total_debit == 0 or total_credit == 0:
                raise ValidationError(
                    _("You cannot save this Invoice because Debit or Credit total is zero.")
                )

    # This function is visible in the actions and is
    # used to create purchase order for bills
    def action_create_po(self):
        in_invoice = self.filtered(lambda x: x.move_type != 'in_invoice')
        if in_invoice:
            raise ValidationError('You can only create purchase order for bills')
        if len(self.mapped('partner_id')) > 1:
            raise ValidationError('Cannot create 1 purchase order for multiple vendors')
        if len(self.mapped('analytic_distribution_id')) > 1:
            raise ValidationError('Multiple Analytic Distribution not allowed')
        else:
            data = {
                'partner_id': self.partner_id.ids[0],
                'analytic_distribution': self.analytic_distribution_id.ids[0],
                'gstin_id': self.gstin_id.ids[0],
                'l10n_in_journal_id': self.journal_id.ids[0],
                'order_line': [(0, 0, {
                    'product_id': rec.product_id.id,
                    'analytic_distribution': rec.analytic_distribution,
                    'date_planned': self.mapped('date')[0],
                    'product_uom': rec.product_uom_id.id,
                    'product_qty': rec.quantity,
                    'price_unit': rec.price_unit,
                    'invoice_lines': [(4, rec.id)]
                }) for rec in self.invoice_line_ids]
            }
            self.env['purchase.order'].create(data)

    @staticmethod
    def convert_num_to_text(value):
        pre = float(value)
        text = ''
        entire_num = int((str(pre).split('.'))[0])
        decimal_num = int((str(pre).split('.'))[1])
        if decimal_num < 10:
            decimal_num = decimal_num * 10
        text += num2words(entire_num, lang='en_IN')
        text += ' Rupees '
        if decimal_num >= 1:
            text += ' and '
            text += num2words(decimal_num, lang='en_IN')
            text += ' Paise only '
        else:
            text += ' only '
        return text.title()

    # Onchange for journal through gstin
    @api.onchange('gstin_id')
    def onchange_gstin_id(self):
        domain = []
        if self.move_type == "out_invoice" or self.move_type == "out_refund":
            print('entereddddddddd...........')
            domain = [('l10n_in_gstin_partner_id.vat', '=', self.gstin_id.vat), ('type', '=', 'sale')]
            journal_ids = self.env['account.journal'].search([("l10n_in_gstin_partner_id", '=', self.gstin_id.id),
                                                              ('type', '=', 'sale')])
            self.journal_id = False
            if len(journal_ids) == 1:
                self.journal_id = journal_ids.id
        elif self.move_type == "in_invoice":
            domain = [('l10n_in_gstin_partner_id.vat', '=', self.gstin_id.vat), ('type', '=', 'purchase')]
            journal_ids = self.env['account.journal'].search([("l10n_in_gstin_partner_id", '=', self.gstin_id.id),
                                                              ('type', '=', 'purchase')])
            self.journal_id = False
            if len(journal_ids) == 1:
                self.journal_id = journal_ids.id
        else:
            if self._context.get('default_custom_payment_type'):
                domain = [('type', '=', 'bank')]
                print('123')
            else:
                print('92')
                domain = [('company_id', '=', self.env.company.id)]

        # domain.append(('id', 'in', self.suitable_journal_ids.ids))
        return {'domain': {'journal_id': domain}}

    # onchange for Place of Supply
    # @api.onchange('journal_id')
    # def onchange_journal_id(self):
    #     if self.journal_id:
    #         self.l10n_in_state_id = self.partner_id.state_id

    # For Tax Calculation
    @api.depends('tax_totals')
    def _compute_gst_amount(self):
        for rec in self:
            rec.igst = 0
            rec.cgst = 0
            rec.sgst = 0
            rec.gst_amount = 0.0
            rec.tds_amount = 0.0
            if not rec.tax_totals:
                continue
            groups_by_subtotal = dict(rec.tax_totals.get('groups_by_subtotal', {}))
            if not groups_by_subtotal:
                continue
            untaxed_amt = groups_by_subtotal.get('Untaxed Amount', ())
            for tax_line in untaxed_amt:
                tax_name = tax_line.get('tax_group_name')
                tax_amt = tax_line.get('tax_group_amount', 0.0)
                if tax_name == 'IGST':
                    rec.igst = tax_amt
                    rec.gst_amount += tax_amt
                elif tax_name == 'CGST':
                    rec.cgst = tax_amt
                    rec.gst_amount += tax_amt
                elif tax_name == 'SGST':
                    rec.sgst = tax_amt
                    rec.gst_amount += tax_amt
                elif tax_name == 'TDS':
                    rec.tds_amount = tax_amt


    @api.onchange('partner_id_line')
    def change_partner(self):
        if self.partner_id_line:
            self.line_ids.partner_id = self.partner_id_line

    @api.depends('amount_total')
    def amt_in_words(self):
        print(self.amount_total)
        self.amt = num2words(self.amount_total, lang='en_IN').replace('and', '').replace('point', 'and').replace(
            'thous', 'thousand')
        print(self.amt)

    def tot_amt_in_words(self, tot_amount):
        return num2words(str(tot_amount), lang='en_IN').replace('and', '').replace('point', 'and').replace(
            'thous', 'thousand')

    def _compute_total_claimable_amt_words(self, amount):
        for rec in self:
            return rec.currency_id.amount_to_text(amount)

    @api.depends('amount_total')
    def _compute_amount_total_words(self):
        for invoice in self:
            invoice.amount_total_words = invoice.currency_id.with_context(lang='en_IN').amount_to_text(
                invoice.amount_total)
        print(invoice.amount_total_words)

    @api.onchange('analytic_distribution_id')
    def apply_analytic(self):
        for line in self.line_ids.filtered(lambda x: x.display_type in ('product', 'payment_term')):
            if self.analytic_distribution_id:
                line.analytic_distribution = {self.analytic_distribution_id.id: 100}

    @api.onchange('analytic_distribution_journal_id')
    def apply_analytic_journal(self):
        for line in self.line_ids:
            if self.analytic_distribution_journal_id:
                line.analytic_distribution = {self.analytic_distribution_journal_id.id: 100}

    # auto populate analytic account in account lines on change of analytic account field
    # CODE FOR ANALYTIC ACCOUNT IN DEBBTORS
    # def write(self, vals):
    #     res = super(AccountMoveInherit, self).write(vals)
    #     if vals.get('analytic_distribution_journal_id') or vals.get('analytic_distribution_id'):
    #         for rec in self:
    #             analytic_list = []
    #             for invoice_line in rec.invoice_line_ids:
    #                 if invoice_line.analytic_distribution:
    #                     if rec.line_ids:
    #                         for line in rec.line_ids:
    #                             if line._origin.analytic_distribution:
    #                                 if line._origin.account_id:
    #                                     if line._origin.account_id.account_type == "income":
    #                                         if line._origin not in analytic_list:
    #                                             analytic_list.append(line._origin)
    #
    #             debtor_analytic_lines = []
    #             for analytic in analytic_list:
    #                 analytic_distribution_dict = {}
    #                 analytic_distribution = analytic.analytic_distribution
    #                 # if len(analytic_distribution) > 1:
    #                 for analytic_dist_key, analytic_dist_value in analytic_distribution.items():
    #                     analytic_credit_with_taxes = analytic_value_with_percent = 0
    #                     analytic_credit_with_taxes = analytic.credit + ((analytic.credit * analytic.tax_ids.amount) / 100)
    #                     analytic_value_with_percent = ((analytic_credit_with_taxes / rec.amount_total) * 100)
    #                     analytic_distribution[analytic_dist_key] = analytic_value_with_percent
    #                     analytic_distribution_dict.update(analytic_distribution)
    #                     if analytic_distribution_dict not in debtor_analytic_lines:
    #                         debtor_analytic_lines.append(analytic_distribution_dict)
    #                 # else:
    #                 #     analytic_distribution_dict.update(analytic_distribution)
    #                 #     if analytic_distribution_dict not in debtor_analytic_lines:
    #                 #         debtor_analytic_lines.append(analytic_distribution_dict)
    #             # count_debtors = list(filter(lambda x: x._origin.account_id.account_type == "asset_receivable", rec.line_ids))
    #
    #             final_analytic_dict_values = {}
    #             for analytic_line in debtor_analytic_lines:
    #                 # for line in rec.line_ids:
    #                 #     if line._origin.account_id.account_type == "asset_receivable":
    #                 for analytic_key, analytic_value in analytic_line.items():
    #                     if analytic_key not in final_analytic_dict_values:
    #                         final_analytic_dict_values[analytic_key] = analytic_value / len(analytic_line)
    #                     else:
    #                         final_analytic_dict_values[analytic_key] += analytic_value / len(analytic_line)
    #
    #             for line in rec.line_ids:
    #                 if line._origin.account_id.account_type == "asset_receivable":
    #                     line._origin['analytic_distribution'] = final_analytic_dict_values
    #     return res

    def cal_hsn_summary(self):
        hsn_code_dict = dict()
        for move_line in self.invoice_line_ids:
            cgst_amt = sgst_amt = igst_amt = 0
            sgst_rate = cgst_rate = igst_rate = 0
            for tax in move_line.tax_ids:
                if 'gst' not in tax.name.lower():
                    continue
                taxes = tax.compute_all(move_line.price_unit, self.currency_id, move_line.quantity,
                                        product=move_line.product_id, partner=move_line.partner_id)
                for tx in taxes['taxes']:
                    name = tx.get('name').lower()
                    if 'cgst' in name:
                        cgst_rate = float(re.findall(r"[-+]?\d*\.\d+|\d+", name)[0])
                        cgst_amt = tx.get('amount', 0)
                    elif 'sgst' in name:
                        sgst_rate = float(re.findall(r"[-+]?\d*\.\d+|\d+", name)[0])
                        sgst_amt = tx.get('amount', 0)
                    elif 'igst' in name:
                        igst_rate = float(re.findall(r"[-+]?\d*\.\d+|\d+", name)[0])
                        igst_amt = tx.get('amount', 0)
            if move_line.hsn_code in hsn_code_dict:
                hsn_code_dict[move_line.hsn_code]['qty'] += move_line.quantity
                hsn_code_dict[move_line.hsn_code]['taxable_value'] += move_line.price_subtotal
                hsn_code_dict[move_line.hsn_code]['cgst_amt'] += cgst_amt
                hsn_code_dict[move_line.hsn_code]['sgst_amt'] += sgst_amt
                hsn_code_dict[move_line.hsn_code]['igst_amt'] += igst_amt
                hsn_code_dict[move_line.hsn_code]['tax_amt'] += move_line.price_total - move_line.price_subtotal
                hsn_code_dict[move_line.hsn_code]['total_amt'] += move_line.price_total
            else:
                hsn_code_dict.update({move_line.hsn_code: {
                    'hsn_code': move_line.hsn_code,
                    'qty': move_line.quantity,
                    'taxable_value': move_line.price_subtotal,
                    'cgst_percent': cgst_rate,
                    'cgst_amt': cgst_amt,
                    'sgst_percent': sgst_rate,
                    'sgst_amt': sgst_amt,
                    'igst_percent': igst_rate,
                    'igst_amt': igst_amt,
                    'tax_amt': move_line.price_total - move_line.price_subtotal,
                    'total_amt': move_line.price_total
                }})
        return hsn_code_dict.values()

    def action_post(self):
        res = super(AccountMoveInherit, self).action_post()
        for invoice in self:
            for line in invoice.invoice_line_ids.filtered(lambda l: l.display_type == 'payment_term' and l.move_type in (
                    'out_invoice', 'out_refund', 'in_invoice', 'in_refund')):
                if self._origin.move_type == 'out_invoice':
                    if not line.analytic_distribution:
                        raise ValidationError('Invoice line for account %s requires an analytic distribution.'
                                              % line.account_id.display_name)
        return res


class AccountMoveLineInherit(models.Model):
    _inherit = 'account.move.line'

    analytic_id = fields.Many2one('account.analytic.plan', related='move_id.analytics_plan_id')
    analytic_dist_id = fields.Many2one('account.analytic.account', related='move_id.analytic_distribution_id')
    # hsn_code = fields.Char(string="HSN Code", compute='compute_hsn_code')
    supplier_date_related = fields.Date(related='move_id.invoice_date', string='Supplier Date')
    untaxed_amt_related = fields.Monetary(related='move_id.amount_untaxed', string='Taxable Amt', store=True)
    gross_amt_related = fields.Monetary(related='move_id.amount_total', string='Gross Amt', store=True)
    ref_bank_no = fields.Char(related="move_id.origin_payment_id.batch_payment_id.ref_bank_no")

    ## ADDED a field to calculate credit value as it will work as a calculator only and applied onchange
    value = fields.Char(string='Value', default='0')
    plan_ids = fields.Many2many('account.analytic.plan')
    analytic_ids = fields.Many2many('account.analytic.account')

    def compute_analytic_data(self):
        for rec in self:
            if rec.analytic_distribution:
                for k, v in rec.analytic_distribution.items():
                    account_id = self.env['account.analytic.account'].search([('id', '=', k)])
                    # rec.analytic_ids = [(4,acc.id) for acc in account_id]
                    rec.plan_ids = [(4, plan.plan_id.id) for plan in account_id]
            else:
                # rec.analytic_ids = False
                rec.plan_ids = False

    @api.onchange('value')
    def compute_prod_avl_qty(self):
        # rec.prod_avl_qty = 0
        for rec in self:
            rec.debit = eval(rec.value)

    # @api.depends('product_id', 'hsn_id')
    # def compute_hsn_code(self):
    #     for rec in self:
    #         rec.hsn_code = rec.product_id.l10n_in_hsn_code or rec.hsn_id.hsnsac_code

    @api.model
    def create(self, vals_list):
        res = super().create(vals_list)
        for rec in res.filtered(lambda x: x.move_type in ('in_invoice', 'out_invoice', 'out_refund', 'in_invoice',
                                                          'in_refund') and x.move_id.analytic_distribution_id and x.display_type in (
                                                  'product', 'payment_term')):
            rec.analytic_distribution = {rec.move_id.analytic_distribution_id.id: 100}
        return res

    # written for changing label in journal item when its Salary Payment
    # @api.model
    # def create(self, values):
    #     res = super(AccountMoveLineInherit, self).create(values)
    #     if res.payment_id:
    #         text = res.name
    #         if res.payment_id.s_check == True:
    #             text = str(text).replace('Vendor', 'Salary')
    #             res.name = text
    #     return res

    # Written for analytic account dependent on analytic plan
    # @api.onchange('analytic_plan')
    # def depart_func_filter(self):
    #     for rec in self:
    #         return {'domain': {'analytics_account': [('plan_id', '=', rec.analytic_plan.id)]}}



# for salary account
class ResSettings(models.TransientModel):
    _inherit = 'res.config.settings'

    bonus_account_id = fields.Many2one(
        comodel_name='account.account',
        string='Bonus Account',
        domain="[('account_type', 'in', ['asset_receivable', 'liability_payable'])]",
        help='Salary Account',
        config_parameter='csspl_india.bonus_account_id')

    salary_account_id = fields.Many2one(
        comodel_name='account.account',
        string='Salary Account',
        domain="[('active', '=', False), ('company_ids', 'in', company_id)]",
        help='Salary Account',
        config_parameter='csspl_india.salary_account_id')

    loans_account_id = fields.Many2one(
        comodel_name='account.account',
        string='Loans Account',
        domain="[('active', '=', False), ('company_ids', 'in', company_id)]",
        help='Loans Account',
        config_parameter='csspl_india.loans_account_id')

    advances_account_id = fields.Many2one(
        comodel_name='account.account',
        string='Advances Account',
        domain="[('active', '=', False), ('company_ids', 'in', company_id)]",
        help='Advances Account',
        config_parameter='csspl_india.advances_account_id')


class ExcelMergeWizard(models.TransientModel):
    _name = 'excel.merge.wizard'

    file_name = fields.Binary(string='Binary')
    summary_data = fields.Char(string='Filename')

    # Report for bank format in xls
    def generate_xlsx_report(self):
        batch_ids = self.env['account.batch.payment'].browse(self.env.context.get('active_ids', []))
        file_name = 'merged_batch_report.xls'
        bold_one = xlwt.easyxf('font: bold 1, height 240; align: vertical center;')

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('TransactionPayeeDetails')
        for col in range(21):
            sheet.col(col).width = 25 * 256  # Setting column width in units of 1/256th of a character width.

        # Header
        header_labels = [
            'Debit_Ac_No', 'Beneficiary_Ac_No', 'Beneficiary', 'Amt', 'PayMode',
            'Date', 'IFSC', 'Payable_Location', 'Print_Location', 'Bene_Mobile_No',
            'Bene_Email_ID', 'Bene_add1', 'Bene_add2', 'Bene_add3', 'Bene_add4',
            'Add_Details_1', 'Add_Details_2', 'Add_Details_3', 'Add_Details_4',
            'Add_Details_5', 'Remarks'
        ]
        for col, label in enumerate(header_labels):
            sheet.write(0, col, label, bold_one)

        row = 1
        col = 0

        for batch_id in batch_ids:
            for line in batch_id.payment_ids:
                if not line.partner_bank_id.bank_id.bic:
                    raise ValidationError(f'Bank/IFSC not set in partner {line.partner_id.name}')
                if not line.batch_payment_id.journal_id.bank_id.bic:
                    raise ValidationError(f'Bank/IFSC not set in Journal {line.partner_id.name}')
                if line.batch_payment_id.journal_id.bank_id.bic[0:4] == line.partner_bank_id.bank_id.bic[0:4]:
                    pay_method = "I"
                elif line.amount > 200000:
                    # current_time = datetime.datetime.now(pytz.timezone('Asia/Kolkata')).time()
                    current_time = fields.Datetime.context_timestamp(self, fields.Datetime.now()).time()
                    if current_time >= time(15, 0):
                        pay_method = "N"
                    else:
                        pay_method = "R"
                else:
                    pay_method = "N"

                partner_name = re.sub(r"[-\d]+", "", line.partner_id.name)

                sheet.write(row, col, line.batch_payment_id.journal_id.bank_account_id.acc_number)
                sheet.write(row, col + 1, line.partner_bank_id.acc_number or ' ')
                sheet.write(row, col + 2, partner_name or ' ')
                sheet.write(row, col + 3, int(line.amount) or ' ')
                sheet.write(row, col + 4, str(pay_method))
                sheet.write(row, col + 5, line.date.strftime("%d-%b-%Y").upper())
                sheet.write(row, col + 6, line.partner_bank_id.bank_id.bic or ' ')
                sheet.write(row, col + 7, ' ')
                sheet.write(row, col + 8, ' ')
                sheet.write(row, col + 9, ' ')
                sheet.write(row, col + 10, ' ')
                sheet.write(row, col + 11, ' ')
                sheet.write(row, col + 12, ' ')
                sheet.write(row, col + 13, ' ')
                sheet.write(row, col + 14, ' ')
                sheet.write(row, col + 15, ' ')
                sheet.write(row, col + 16, ' ')
                sheet.write(row, col + 17, ' ')
                sheet.write(row, col + 18, ' ')
                sheet.write(row, col + 19, ' ')
                sheet.write(row, col + 20, line.narration)

                row += 1

        # Save the workbook to a file or return it as a response to a web request.
        fp = BytesIO()
        workbook.save(fp)
        self.write({'file_name': base64.b64encode(fp.getvalue()), 'summary_data': file_name})
        fp.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'excel.merge.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'target': 'new',
        }

class AccountBatchPaymentInherit(models.Model):
    _inherit = "account.batch.payment"

    rej_reason = fields.Text('Reject Reason', tracking=True)
    subject_line = fields.Char('Subject Line')
    attachment_batch = fields.Binary(string='Mail Attachment', attachment=True)
    working_attachment = fields.Binary(string='Working Attachment', attachment=True, tracking=True)
    pdf_drawing_name = fields.Char(tracking=True)
    pdf_attachment_name = fields.Char(tracking=True)
    to_check = fields.Boolean("To Check", tracking=True, copy=False, default=False)
    is_checked = fields.Boolean("Checked", tracking=True, copy=False, default=False)
    checker_id = fields.Many2one('res.users', "Checker")
    is_reject = fields.Boolean()
    # temporarily commented
    state = fields.Selection([
        ('draft', 'New'),
        # ('submitted', 'Submitted'),
        # ('approved', 'Approved'),
        # ('reject', 'Rejected'),
        # ('management_approved', 'Management Approved'),
        ('bank_upload', 'Bank Upload'),
        ('pending_transfer', 'Pending For Transfer'),
        ('sent', 'Transfered'),
        ('reconciled', 'Reconciled'),
    ], store=True, compute='_compute_state', default='draft', track_visibility='onchange')

    ref_bank_no = fields.Char('Reference Bank No', tracking=True)
    analytics_account_id = fields.Many2one('account.analytic.account', string='Analytics Account')
    bank_date = fields.Date('Transfer pending Date', tracking=True)
    is_management_approved = fields.Boolean('is_management_approved')
    is_checked_approve = fields.Boolean('Is Checked')
    cancelled_payments_ids = fields.Many2many('account.payment', string='Payment Ids')
    cancelled_order_count = fields.Integer(string='Cancelled Payments', compute='compute_cancelled_count',
                                           copy=False)
    today_date = fields.Datetime.now()
    accepted = fields.Integer('Accepted', compute='compute_calculate_payments')
    rejected = fields.Integer('Rejected', compute='compute_calculate_payments')
    total_count = fields.Integer('Total Count', compute='compute_calculate_payments')
    total_payment = fields.Float('Total Payment', compute='compute_calculate_payments')
    total_accepted = fields.Float('Total Accepted', compute='compute_calculate_payments')
    total_rejected = fields.Float('Total Rejected', compute='compute_calculate_payments')
    analytics_plans_batch_id = fields.Many2one('account.analytic.plan', string='Analytics Plan')

    ## ADDED BY KAJAL
    amt_total_value = fields.Float(string='Total Amount')
    approved_date = fields.Datetime(string='Approved Date')
    is_amt_bool = fields.Boolean(
        string="Custom Boolean", default=False,
        compute='_compute_is_amt',
    )
    prioritys = fields.Boolean()

    def write(self, vals):
        if self.payment_ids:
            data = self.payment_ids[0]
            value = data.analytics_account_id.id
            vals.update({'analytics_account_id': value})
        else:
            pass
        res = super(AccountBatchPaymentInherit, self).write(vals)
        if 'bank_date' in vals:
            self.set_month_in_payment()
        return res

    @api.onchange('payment_ids')
    def _onchange_payment_ids_validate_analytics(self):
        for batch in self:
            payments = batch.payment_ids
            if payments:
                first_payment = payments[0]
                first_account = first_payment.analytics_account_id
                if first_account:
                    batch.analytics_account_id = first_account.id
                else:
                    batch.analytics_account_id = False
                # first_plan = first_payment.analytics_plan_id

                # inconsistent_account = any(
                #     p.analytics_account_id != first_account for p in payments
                # )
                # inconsistent_plan = any(
                #     p.analytics_plan_id != first_plan for p in payments
                # )
                #
                # if inconsistent_account or inconsistent_plan:
                #     raise ValidationError(
                #         f"All Batch must have the same Analytics Account ({first_account.name}) and Analytics Plan ({first_plan.name}).")
                # else:
                #     pass
            else:
                continue

    # amt_total = fields.Char(string='Total Amount')

    @api.depends('amount')
    def _compute_is_amt(self):
        for rec in self:
            rec.is_amt_bool = True
            if rec.amount:
                rec.amt_total_value = rec.amount
            else:
                rec.amt_total_value = 0.0

    ##

    # @api.onchange('bank_date')
    def set_month_in_payment(self):
        if self.payment_ids:
            payment_to_update = self.payment_ids.filtered(lambda x: x.payment_for in ('other', False))
            if self.bank_date:
                payment_month = self.env['payments.month'].search(
                    [('name', '=', self.bank_date.strftime('%b-%y'))], limit=1)
                payment_to_update.payment_month_id = payment_month.id if payment_month else False
            else:
                payment_to_update.payment_month_id = False

    @api.onchange('journal_id')
    def change_journal_in_payments(self):
        if self.payment_ids.filtered(lambda x: x.state == 'posted'):
            raise ValidationError("You cannot change the journal as there are posted entries.")
        if self.cancelled_payments_ids.filtered(lambda x: x.state == 'posted'):
            raise ValidationError("Rejected Payments are posted, cannot change the journal.")
        if self.journal_id:
            for rec in (self.payment_ids + self.cancelled_payments_ids):
                rec._origin.journal_id = self.journal_id.id

    # Added for Merging Custom format
    def generate_excel_merge_report(self):
        return {
            'name': 'Merge Excel Format',
            'res_model': 'excel.merge.wizard',
            'view_mode': 'form',
            'type': 'ir.actions.act_window',
            'target': 'new',
            'context': {'active_id': self.ids},

        }
    # # temporarily commmented
    # def resend_for_approval(self):
    #     for rec in self.cancelled_payments_ids:
    #         rec.batch_payment_id = self.id
    #         rec.action_draft()
    #         self.cancelled_payments_ids = [(3, rec.id)]
    #     self.state = 'submitted'

    # Report for bank format in xlsx For Attaching in Mail
    def generate_xlsx_mail_report(self):
        workbook = Workbook()
        sheet = workbook.create_sheet('TransactionPayeeDetails')

        sheet.cell(row=1, column=1, value='Bank')
        sheet.cell(row=1, column=2, value='Date')
        sheet.cell(row=1, column=3, value='Reference')
        sheet.cell(row=1, column=4, value='Batch State')
        sheet.cell(row=1, column=5, value='Accepted')
        sheet.cell(row=1, column=6, value='Rejected')
        sheet.cell(row=1, column=7, value='Total Count')
        sheet.cell(row=1, column=8, value='Payments/Customer/Vendor')
        sheet.cell(row=1, column=9, value='Payments/Date')
        sheet.cell(row=1, column=10, value='Payments/Amount')
        sheet.cell(row=1, column=11, value='Payments/Analytic Account')
        sheet.cell(row=1, column=12, value='Payments/Month')
        sheet.cell(row=1, column=13, value='Payments/Bank')
        sheet.cell(row=1, column=14, value='Payments/Account Number')
        sheet.cell(row=1, column=15, value='Payments/Bank Identifier Code')
        sheet.cell(row=1, column=16, value='Payments/Reject Reason')
        sheet.cell(row=1, column=17, value='Payments/Description')
        sheet.cell(row=1, column=18, value='Payments/State')

        row = 2
        total_payments_search = ''
        for line in self.payment_ids:
            total_payments_search = self.env['account.payment'].search(
                [('source_doc', '=', line.batch_payment_id.name)])

            sheet.cell(row=row, column=1, value=line.batch_payment_id.journal_id.name)
            sheet.cell(row=row, column=2, value=line.batch_payment_id.date or ' ')
            sheet.cell(row=row, column=3, value=line.batch_payment_id.name or ' ')
            sheet.cell(row=row, column=4, value=line.batch_payment_id.state or ' ')
            sheet.cell(row=row, column=5, value=line.batch_payment_id.accepted or ' ')
            sheet.cell(row=row, column=6, value=len(line.batch_payment_id.cancelled_payments_ids) or ' ')
            sheet.cell(row=row, column=7, value=line.batch_payment_id.total_count or ' ')

        for i in total_payments_search:
            sheet.cell(row=row, column=8, value=i.partner_id.name or ' ')
            sheet.cell(row=row, column=9, value=i.date or '')
            sheet.cell(row=row, column=10, value=i.amount_signed or '')
            sheet.cell(row=row, column=11, value=i.analytics_account_id.name or '')
            sheet.cell(row=row, column=12, value=i.month or '')
            sheet.cell(row=row, column=13, value=i.partner_bank_id.bank_id.name or '')
            sheet.cell(row=row, column=14, value=i.bank_acc_related or '')
            sheet.cell(row=row, column=15, value=i.partner_bank_id.bank_id.bic or '')
            sheet.cell(row=row, column=16, value=i.reason_text or '')
            sheet.cell(row=row, column=17, value=i.narration or '')
            sheet.cell(row=row, column=18, value=i.state or '')
            row += 1

        # Create a BytesIO object to hold the file data
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        return excel_file

    # Function for calculating summary of payments:
    def compute_calculate_payments(self):
        for rec in self:
            rec.accepted = len(rec.payment_ids.ids)
            rec.rejected = len(rec.cancelled_payments_ids.ids)
            rec.total_count = len(rec.payment_ids.ids) + len(rec.cancelled_payments_ids.ids)
            rec.total_payment = sum(rec.payment_ids.mapped('amount_signed')) + sum(
                rec.cancelled_payments_ids.mapped('amount_signed'))
            rec.total_accepted = sum(rec.payment_ids.mapped('amount_signed'))
            rec.total_rejected = sum(rec.cancelled_payments_ids.mapped('amount_signed'))

    # Download button for excel report
    # def download_payments(self):
    #     return self.env.ref('csspl_india.payment_report_xlsx').report_action(self)

    # Added for Adding cancelled payments
    def open_cancelled_payments(self):
        action = self.env.ref('account.action_account_payments_payable').sudo().read()[0]
        if self.cancelled_payments_ids:
            action['domain'] = [('id', 'in', self.cancelled_payments_ids.ids)]
        else:
            action = {'type': 'ir.actions.act_window_close'}
        return action

    def compute_cancelled_count(self):
        for rec in self:
            rec.cancelled_order_count = len(rec.cancelled_payments_ids.ids)

    # Inherited for batch payment status
    @api.depends('payment_ids.move_id.is_move_sent', 'payment_ids.is_sent', 'payment_ids.is_matched', 'to_check',
                 'is_checked', 'is_reject')
    def _compute_state(self):
        for batch in self:
            test = all(pay.is_matched for pay in batch.payment_ids)
            test2 = all(pay.is_sent for pay in batch.payment_ids)
            test3 = batch.payment_ids.filtered(lambda x: x.is_matched is False)
            if batch.payment_ids and all(pay.is_matched and pay.is_sent for pay in batch.payment_ids):
                batch.state = 'reconciled'
            elif batch.payment_ids and all(pay.is_sent for pay in batch.payment_ids):
                batch.state = 'sent'
            # temporarily commented
            # elif batch.to_check == True:
            #     batch.state = 'submitted'
            # elif batch.is_checked == True:
            #     batch.state = 'approved'
            # elif batch.is_reject == True:
            #     batch.state = 'reject'

    # temporarily commented
    # def send_for_checking(self):
    #     self.state_mail_submitted()
    #     return {
    #         'name': _('Send for Checker'),
    #         'res_model': 'send.checker',
    #         'view_mode': 'form',
    #         'target': 'new',
    #         'payment_id': self.id,
    #         'type': 'ir.actions.act_window'
    #     }
    # # temporarily commmented
    # def button_set_checked(self):
    #     for payment in self:
    #         payment.to_check = False
    #         payment.is_checked = True
    #         payment.is_checked_approve = True
    #         payment.state_mail_approved()

    # # temporarily commmented
    # def reject(self):
    #     for move in self:
    #         if not move.rej_reason:
    #             raise ValidationError('Kindly Mention the Reject Reason')
    #         move.is_reject = True
    #         move.is_checked = False
    #         move.to_check = False
    #
    #         for payment in move.payment_ids:
    #             if not payment.payment_month_id:
    #                 month = self.env['payments.month'].search([('name', '=', payment.month)], limit=1)
    #                 if month:
    #                     payment.payment_month_id = month.id
    #             payment.reason_text = move.rej_reason
    #             payment.batch_payment_id.cancelled_payments_ids = [(4, payment.id)]
    #             payment.batch_payment_id = False
    #             payment.action_cancel()
    #
    #         template = self.env.ref('csspl_india.mail_template_of_batch_rejection')
    #         if move.payment_ids:
    #             requestors = move.payment_ids.mapped('request_by')
    #         else:
    #             requestors = move.cancelled_payments_ids.mapped('request_by')
    #         email_list = list(filter(None, requestors.mapped('email')))
    #         email_list.extend([
    #             'ayadav@cssindia.in',
    #             'singhpooja@cssindia.in',
    #         ])
    #         emails = ','.join(set(email_list))
    #         if emails:
    #             template.email_to = emails
    #         else:
    #             template.email_to = False
    #         template.send_mail(move.id, force_send=True)
    #         move.message_post(body="Mail has been sent to: %s" % (emails or "No recipients"))

    def action_draft(self):
        for payment in self:
            payment.state = 'draft'
            payment.to_check = False
            payment.is_checked = False

    def validate_batch_button(self):
        res = super(AccountBatchPaymentInherit, self).validate_batch_button()
        if not self.bank_date or not self.ref_bank_no:
            raise ValidationError('Set Reference No/Transfer Date first')
        for payment in self.filtered(lambda x: x.state == 'draft'):
            # payment.state_mail_transfered()
            for line in payment.payment_ids:
                line.state = 'paid'
        return res

    def batch_validate_action(self):
        for rec in self:
            rec.validate_batch_button()

    # For Sending mail to followers
    def state_mail_submitted(self):
        mail_values = {}
        excel_data = self.generate_xlsx_mail_report()

        for rec in self:
            if rec._origin.message_follower_ids:
                to = ""
                body = ""
                for follower in rec._origin.message_follower_ids:
                    to += str(follower.email) + ","
                    body = f"Dear Concerned,<br/><br/>" \
                           f" With Reference to the request you raised, " \
                           f"Batch Payment  <b>({rec.name})</b> has been created for the amount of <b>Rs {format_decimal(abs(rec.total_accepted), locale='en_IN')}</b> on <b>{rec.date.strftime('%d-%B-%Y')}</b><br/><br/><br/>" \
                           f"Attached Excel File for Reference"

                # Attach the Excel file to the email
                attachment_vals = {
                    'name': 'Payment_Data.xlsx',
                    'datas': base64.b64encode(excel_data.read()),
                    'store_fname': 'Payment_Data.xlsx',
                }

                # Compose email

                mail_values = {
                    'email_to': to,
                    'subject': rec.subject_line,
                    'body_html': body,
                    'state': 'outgoing',
                    'attachment_ids': [(0, 0, attachment_vals)],
                }
            sendmail = self.env['mail.mail'].sudo().create(mail_values)
            sendmail.send()

    def state_mail_approved(self):
        mail_values = {}
        excel_data = self.generate_xlsx_mail_report()

        for rec in self:
            if rec._origin.message_follower_ids:
                to = ""
                body = ""
                for follower in rec._origin.message_follower_ids:
                    to += str(follower.email) + ","
                    body = f"Dear Concerned,<br/><br/>" \
                           f" With Reference to the request you raised, Your Batch <b>({rec.name})</b>" \
                           f" for the amount of <br/><b>Rs {format_decimal(abs(rec.total_accepted), locale='en_IN')} /-</b> is approved by {rec.env.user.name} on <b>{rec.date.strftime('%d-%B-%Y')}</b><br/><br/><br/>" \
                           f"Attached Excel File For Reference<br/><br/>" \
                           f"""

                                <html><body>
                                <div style="display:table;">

                                    <table style="width: 100%;border: solid 1px black;">
                                    <tr><td style="border: solid 1px black;">Total Payment</td>
                                        <td style="border: solid 1px black;">Accepted Payment</td>
                                        <td style="border: solid 1px black;">Rejected Payment</td></tr>
                                    <tr><td style="border: solid 1px black;text-align: right">{rec.total_count}</td>
                                    <td style="border: solid 1px black;text-align: right">{rec.accepted}</td>
                                    <td style="border: solid 1px black;text-align: right">{rec.rejected}</td></tr>

                                    </table>
                                </div>


                                </body></html>
                                """

                # Attach the Excel file to the email
                attachment_vals = {
                    'name': 'Payment_Data.xlsx',
                    'datas': base64.b64encode(excel_data.read()),
                    'store_fname': 'Payment_Data.xlsx',
                }

                # Compose email

                mail_values = {
                    'email_to': to,
                    'subject': rec.subject_line,
                    'body_html': body,
                    'state': 'outgoing',
                    'attachment_ids': [(0, 0, attachment_vals)],
                }
            sendmail = self.env['mail.mail'].sudo().create(mail_values)
            sendmail.send()

    def state_mail_transfered(self):
        mail_values = {}
        excel_data = self.generate_xlsx_mail_report()

        for rec in self:
            if rec._origin.message_follower_ids:
                to = ""
                body = ""
                for follower in rec._origin.message_follower_ids:
                    to += str(follower.email) + ","
                    body = f"Dear Concerned,<br/><br/>" \
                           f" With Reference to the request you raised, Your Batch <b>{rec.name}</b> has been Transfered " \
                           f" for the amount of <br/><b>Rs {format_decimal(abs(rec.total_accepted), locale='en_IN')} /-</b> on <b>{rec.bank_date.strftime('%d-%B-%Y')}</b><br/><br/><br/>" \
                           f"Attached Excel File For Reference<br/><br/>" \
                           f"""

                                            <html><body>
                                            <div style="display:table;">

                                                <table style="width: 100%;border: solid 1px black;">
                                                <tr><td style="border: solid 1px black;">Total Payment</td>
                                                    <td style="border: solid 1px black;">Accepted Payment</td>
                                                    <td style="border: solid 1px black;">Rejected Payment</td></tr>
                                                <tr><td style="border: solid 1px black;text-align: right">{rec.total_count}</td>
                                                <td style="border: solid 1px black;text-align: right">{rec.accepted}</td>
                                                <td style="border: solid 1px black;text-align: right">{rec.rejected}</td></tr>

                                                </table>
                                            </div>


                                            </body></html>
                                            """

                # Attach the Excel file to the email
                attachment_vals = {
                    'name': 'Payment_Data.xlsx',
                    'datas': base64.b64encode(excel_data.read()),
                    'store_fname': 'Payment_Data.xlsx',
                }

                # Compose email

                mail_values = {
                    'email_to': to,
                    'subject': rec.subject_line,
                    'body_html': body,
                    'state': 'outgoing',
                    'attachment_ids': [(0, 0, attachment_vals)],
                }
            sendmail = self.env['mail.mail'].sudo().create(mail_values)
            sendmail.send()

    def reject_batch_mail(self):
        mail_values = {}
        # excel_data = self.generate_xlsx_mail_report()

        for rec in self:
            if rec._origin.message_follower_ids:
                to = ""
                body = ""
                for follower in rec._origin.message_follower_ids:
                    to += str(follower.email) + ","
                    body = f"Dear Concerned,<br/><br/>" \
                           f" With Reference to the request you raised, Your Batch <b>{rec.name}</b> has been Rejected  " \
                           f" for the amount of <br/><b>Rs {format_decimal(abs(rec.total_accepted), locale='en_IN')} /-</b> on <b>{rec.date.strftime('%d-%B-%Y')}</b> due to the reason {rec.rej_reason}<br/><br/><br/>" \
                           f"""

                                <html><body>
                                <div style="display:table;">

                                    <table style="width: 100%;border: solid 1px black;">
                                    <tr><td style="border: solid 1px black;">Total Payment</td>
                                        <td style="border: solid 1px black;">Accepted Payment</td>
                                        <td style="border: solid 1px black;">Rejected Payment</td></tr>
                                    <tr><td style="border: solid 1px black;text-align: right">{rec.total_count}</td>
                                    <td style="border: solid 1px black;text-align: right"> 0 </td>
                                    <td style="border: solid 1px black;text-align: right">{rec.rejected}</td></tr>

                                    </table>
                                </div>


                                </body></html>
                                """

                # Attach the Excel file to the email
                # attachment_vals = {
                #     'name': 'Payment_Data.xlsx',
                #     'datas': base64.b64encode(excel_data.read()),
                #     'store_fname': 'Payment_Data.xlsx',
                # }

                # Compose email

                mail_values = {
                    'email_to': to,
                    'subject': rec.subject_line,
                    'body_html': body,
                    'state': 'outgoing',
                    # 'attachment_ids': [(0, 0, attachment_vals)],
                }
            sendmail = self.env['mail.mail'].sudo().create(mail_values)
            sendmail.send()

    # def reject_mail(self):
    #     if self._origin.message_follower_ids:
    #         to = ""
    #         for follower in self._origin.message_follower_ids:
    #             to += str(follower.email) + ","
    #             for line in self.payment_ids:
    #                 body = f" Payment({line.narration}) of, <b>{line.partner_id.name}</b> in Batch Payment {self.name}, with account no {line.bank_acc_related}, {line.partner_bank_id.bank_id.bic} is Rejected , for reason {line.reason_text}"
    #
    #         mail_values = {
    #             'email_to': to,
    #             'subject': self.subject_line + '(Rejected Payment)',
    #             'body_html': body,
    #             'state': 'outgoing',
    #         }
    #
    #         sendmail = self.env['mail.mail'].sudo().create(mail_values)
    #         sendmail.send()

    # temporarily commmented
    # def management_approve(self):
    #     for rec in self:
    #         note = "Batch payment %s was Approved on %s by %s" % (
    #             rec.name, datetime.now(), rec.activity_user_id.name)
    #         rec.message_post(body=note)
    #         rec.is_management_approved = True
    #         rec.approved_date = datetime.now()
    #         rec.is_checked = False
    #         rec.state = 'management_approved'

    # temporarily commented
    # def batch_management_approval(self):
    #     for rec in self:
    #         if rec.state == 'approved':
    #             rec.management_approve()

    def pending_transfer(self):
        for rec in self:
            if not rec.ref_bank_no:
                raise ValidationError('Please Enter Reference Bank No')
            else:
                rec.state = 'pending_transfer'

    def bank_upload(self):
        for rec in self:
            rec.is_checked = False
            rec.state = 'bank_upload'

    def update_bank_reference(self):
        return {
            'name': 'Update Bank Reference',
            'res_model': 'bank.ref',
            'view_mode': 'form',
            'type': 'ir.actions.act_window',
            'target': 'new',
            'context': {'active_id': self.ids},
        }

    def action_download_excel(self):
        self.ensure_one()

        batch = self
        output = io.BytesIO()

        # Create workbook & worksheet in memory
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Batch Payment Report')

        bold = workbook.add_format({'bold': True})
        money = workbook.add_format({'num_format': '#,##0.00'})

        # Headers
        headers = [
            'Batch Date', 'Batch Name', 'Analytic Account', 'Journal',
            'Batch Amount', 'Batch State',
            'Partner Name', 'Payment Date', 'Payment Amount',
            'Partner Bank', 'Bank Account Number', 'Bank BIC'
        ]
        for col, header in enumerate(headers):
            sheet.write(0, col, header, bold)

        all_payments = batch.payment_ids | batch.cancelled_payments_ids

        row = 1
        for payment in all_payments:
            sheet.write(row, 0, batch.date or '')
            sheet.write(row, 1, batch.name or '')
            sheet.write(row, 2, batch.analytics_account_id.name or '')
            sheet.write(row, 3, batch.journal_id.name or '')
            sheet.write(row, 4, batch.amount or 0, money)
            sheet.write(row, 5, batch.state or '')

            sheet.write(row, 6, payment.partner_id.name or '')
            sheet.write(row, 7, str(payment.date) or '')
            sheet.write(row, 8, payment.amount or 0, money)
            sheet.write(row, 9, payment.partner_bank_id.bank_id.name or '')
            sheet.write(row, 10, payment.partner_bank_id.acc_number or '')
            sheet.write(row, 11, payment.partner_bank_id.bic or '')

            row += 1

        workbook.close()
        output.seek(0)
        excel_data = output.read()
        output.close()
        excel_base64 = base64.b64encode(excel_data)
        filename = f"Batch_Payment_{batch.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': excel_base64,
            'store_fname': filename,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Return directly as download (no DB write)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }


class InheritAccountingPayment(models.Model):
    _inherit = "account.payment"

    utr_no = fields.Char()
    is_move_sent = fields.Boolean()
    # batch_payment_id = fields.Many2one('account.batch.payment', ondelete='set null', copy=False,
    #                                    compute="_compute_batch_payment_id", store=True, readonly=False, tracking=True)

    # partner_id = fields.Many2one(
    #     comodel_name='res.partner',
    #     string="Customer/Vendor",
    #     store=True, readonly=False, ondelete='restrict',
    #     compute='_compute_partner_id',
    #     tracking=True,
    #     check_company=True)

    transfer_date = fields.Date(compute="set_transfer_date", store=True, string="Transfer Date")
    cheque_number = fields.Integer(string="Cheque No")
    cheque_number1 = fields.Char(string="Cheque No")
    cheque_date = fields.Date(string="Cheque Date")
    payment_for = fields.Selection([('salary', 'Salary'), ('rent', 'Rent'), ('advance', 'Advance'), ('loan', 'Loan'),
                                    ('expense', 'Expenses'), ('other', 'Other'), ('bonus', 'Bonus')], default='other')
    s_check = fields.Boolean(
        string='Salary Payment',
        tracking=True, copy=False, default=False,
        help="Set Default Salary account here",
    )
    c_check = fields.Boolean(
        string='Conveyence Booking',
        tracking=True, copy=False, default=False,
        help="Set Conveyence Booking account here",
    )

    advances_check = fields.Boolean(
        string='Advances',
        tracking=True, copy=False, default=False,
        help="Set Default Advances account here",
    )
    loans_check = fields.Boolean(
        string='Loans',
        tracking=True, copy=False, default=False,
        help="Set Loans account here",
    )
    analytics_account_id = fields.Many2one('account.analytic.account', string='Analytics Account')
    analytics_plan_id = fields.Many2one('account.analytic.plan', string='Analytics Plan')

    request_by = fields.Many2one('res.partner', 'REQUEST BY')
    request_received_date = fields.Date('REQUEST RECEIVED DATE')
    css_local_branch = fields.Char('CSS LOCAL BRANCH')
    css_state = fields.Char('State')
    sup_name = fields.Text('SUP NAME')
    remark = fields.Text('REMARK')
    customer = fields.Many2one('res.partner', 'CUSTOMER')
    customer_ch = fields.Char('CUSTOMER')
    # bank_id = fields.Many2one('res.bank')
    bank = fields.Char(string="BANK")
    atm_id = fields.Char('ATM ID')
    month = fields.Char('MONTH', tracking=True)
    mail_by = fields.Many2one('res.users', 'Mail By')
    requester_mail_id = fields.Char('Requester Mail id')
    unique_id = fields.Char('Unique Id')  # not used
    reason_text = fields.Text('Reject Reason')
    narration = fields.Text('Narration')
    bank_acc_related = fields.Char(related='partner_bank_id.acc_number', string='Account Number')
    source_doc = fields.Char('Source Doc')
    today_date = fields.Datetime.now()
    # contigency = fields.Float('Contigency', compute="compute_contigency_value", store=True)

    @api.onchange('ref')
    def onchange_of_ref(self):
        for rec in self:
            rec.payment_reference = False
            if rec.ref:
                rec.payment_reference = rec.ref

    # @api.depends('amount', 'move_id', 'move_id.reversal_move_ids', 'move_id.reversal_move_ids.state')
    # def compute_contigency_value(self):
    #     for rec in self:
    #         filter_reversal = rec.move_id.reversal_move_ids.filtered(lambda x: x.state == 'posted')
    #         if rec.move_id and filter_reversal:
    #             rec.contigency = rec.amount - sum(filter_reversal.mapped('amount'))
    #         else:
    #             rec.contigency = rec.amount

    # For analytic distribution from payment
    def action_post(self):
        res = super(InheritAccountingPayment, self).action_post()
        if self.analytics_account_id:
            self.move_id.line_ids.analytic_distribution = {self.analytics_account_id.id: 100}
        # if (self.partner_bank_id.allow_out_payment is False) and (self.payment_type == 'outbound') and (
        #         self.is_internal_transfer == False) and (self.partner_id.cust_partner_type != 'compliance'):
        #     raise ValidationError('Payment receipt is not allowed for this contact.')
        # return res

    @api.depends('date', 'batch_payment_id', 'batch_payment_id.bank_date')
    def set_transfer_date(self):
        for rec in self:
            rec.transfer_date = None
            if rec.batch_payment_id and rec.batch_payment_id.bank_date:
                rec.transfer_date = rec.batch_payment_id.bank_date
            elif not rec.batch_payment_id:
                rec.transfer_date = False

    # Inherited for sal payment
    # @api.depends('payment_for', 'journal_id', 'partner_id', 'partner_type', 'is_internal_transfer',
    #              'destination_journal_id')
    # def _compute_destination_account_id(self):
    #     self.destination_account_id = False
    #     for pay in self:
    #         if pay.is_internal_transfer:
    #             pay.destination_account_id = pay.destination_journal_id.company_id.transfer_account_id
    #
    #         elif pay.payment_for == 'bonus':
    #             pay.destination_account_id = int(
    #                 self.env['ir.config_parameter'].sudo().get_param('csspl_india.bonus_account_id')) or False
    #         elif pay.payment_for == 'salary':
    #             pay.destination_account_id = int(
    #                 self.env['ir.config_parameter'].sudo().get_param('csspl_india.salary_account_id')) or False
    #         elif pay.payment_for == 'loan':
    #             pay.destination_account_id = int(
    #                 self.env['ir.config_parameter'].sudo().get_param('csspl_india.loans_account_id')) or False
    #         elif pay.payment_for == 'advance':
    #             pay.destination_account_id = int(
    #                 self.env['ir.config_parameter'].sudo().get_param('csspl_india.advances_account_id')) or False
    #
    #         elif pay.partner_type == 'customer':
    #             # Receive money from invoice or send money to refund it.
    #             if pay.partner_id:
    #                 pay.destination_account_id = pay.partner_id.with_company(
    #                     pay.company_id).property_account_receivable_id
    #             else:
    #                 pay.destination_account_id = self.env['account.account'].search([
    #                     ('company_id', '=', pay.company_id.id),
    #                     ('account_type', '=', 'asset_receivable'),
    #                     ('deprecated', '=', False),
    #                 ], limit=1)
    #         elif pay.partner_type == 'supplier':
    #             # Send money to pay a bill or receive money to refund it.
    #             if pay.partner_id:
    #                 pay.destination_account_id = pay.partner_id.with_company(pay.company_id).property_account_payable_id
    #             else:
    #                 pay.destination_account_id = self.env['account.account'].search([
    #                     ('company_id', '=', pay.company_id.id),
    #                     ('account_type', '=', 'liability_payable'),
    #                     ('deprecated', '=', False),
    #                 ], limit=1)

    def reject_payment(self):
        if self.batch_payment_id:
            if not self.reason_text:
                return {
                    'name': 'Select Reason',
                    'res_model': 'reason.wiz',
                    'view_mode': 'form',
                    'type': 'ir.actions.act_window',
                    'target': 'new',
                }
            # self.batch_payment_id.reject_mail()
            self.batch_payment_id.cancelled_payments_ids = [(4, self.id)]
            self.batch_payment_id = False
            self.action_cancel()

    def action_cancel(self):
        res = super().action_cancel()
        for rec in self:
            if rec.batch_payment_id and rec.batch_payment_id.state == 'sent':
                raise ValidationError(
                    f"Cannot cancel the entry as the Batch Payment {rec.batch_payment_id.name} is in {rec.batch_payment_id.state} state")
        return res


class ChartAccount(models.Model):
    _inherit = 'account.account'

    is_tds_ledger = fields.Boolean(string="TDS Ledger")


class AccountInvoiceReport(models.Model):
    _inherit = "account.invoice.report"

    service_date = fields.Date('Service Date')

    #
    #     @api.model
    # def _select(self):
    #     return '''
    #         SELECT
    #             line.id,
    #             line.move_id,
    #             line.product_id,
    #             line.account_id,
    #             line.journal_id,
    #             line.company_id,
    #             line.company_currency_id,
    #             line.partner_id AS commercial_partner_id,
    #             account.account_type AS user_type,
    #             move.state,
    #             move.move_type,
    #             move.partner_id,
    #             move.invoice_user_id,
    #             move.fiscal_position_id,
    #             move.payment_state,
    #             move.invoice_date,
    #             move.service_date,
    #             move.invoice_date_due,
    #             uom_template.id                                             AS product_uom_id,
    #             template.categ_id                                           AS product_categ_id,
    #             line.quantity / NULLIF(COALESCE(uom_line.factor, 1) / COALESCE(uom_template.factor, 1), 0.0) * (CASE WHEN move.move_type IN ('in_invoice','out_refund','in_receipt') THEN -1 ELSE 1 END)
    #                                                                         AS quantity,
    #             -line.balance * currency_table.rate                         AS price_subtotal,
    #             line.price_total * (CASE WHEN move.move_type IN ('in_invoice','out_refund','in_receipt') THEN -1 ELSE 1 END)
    #                                                                         AS price_total,
    #             -COALESCE(
    #                -- Average line price
    #                (line.balance / NULLIF(line.quantity, 0.0)) * (CASE WHEN move.move_type IN ('in_invoice','out_refund','in_receipt') THEN -1 ELSE 1 END)
    #                -- convert to template uom
    #                * (NULLIF(COALESCE(uom_line.factor, 1), 0.0) / NULLIF(COALESCE(uom_template.factor, 1), 0.0)),
    #                0.0) * currency_table.rate                               AS price_average,
    #             COALESCE(partner.country_id, commercial_partner.country_id) AS country_id,
    #             line.currency_id                                            AS currency_id
    #     ''' + ", move.team_id as team_id"


class CustomExcel(models.TransientModel):
    _name = 'custom.excel.class'
    _rec_name = 'summary_data'

    file_name = fields.Binary(string='Binary')
    summary_data = fields.Char(string='Filename')

    bank_select = fields.Selection([('axis', 'Axis Bank'),
                                    ('icici', 'ICICI Bank')], string='Bank')

    def generate_xlsx_report(self):
        if self.bank_select == 'axis':
            return self.generate_xlsx_report_icici()
        if self.bank_select == 'icici':
            return self.generate_xlsx_report_axis()

    def generate_xlsx_report_axis(self):

        batch_id = self.env['account.batch.payment'].browse(self.env.context['active_id'])
        file_name = f"{batch_id.name}_{fields.Date.today().strftime('%d%B%Y')}.xls"
        # file_name = (batch_id.name) + '_' + (batch_id.today_date.strftime('%d%B%Y')) + '.xls'
        bold_one = xlwt.easyxf('font: bold 1, height 240; align: vertical center;')
        # title_date_format = xlwt.easyxf(num_format_str='d-MMM-yyyy')

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('TransactionPayeeDetails')
        for col in range(21):
            sheet.col(col).width = 25 * 256  # Setting column width in units of 1/256th of a character width.

        # Header
        sheet.write(0, 0, 'Debit_Ac_No', bold_one)
        sheet.write(0, 1, 'Beneficiary_Ac_No', bold_one)
        sheet.write(0, 2, 'Beneficiary', bold_one)
        sheet.write(0, 3, 'Amt', bold_one)
        sheet.write(0, 4, 'PayMode', bold_one)
        sheet.write(0, 5, 'Date', bold_one)
        sheet.write(0, 6, 'IFSC', bold_one)
        sheet.write(0, 7, 'Payable_Location', bold_one)
        sheet.write(0, 8, 'Print_Location', bold_one)
        sheet.write(0, 9, 'Bene_Mobile_No', bold_one)
        sheet.write(0, 10, 'Bene_Email_ID', bold_one)
        sheet.write(0, 11, 'Bene_add1', bold_one)
        sheet.write(0, 12, 'Bene_add2', bold_one)
        sheet.write(0, 13, 'Bene_add3', bold_one)
        sheet.write(0, 14, 'Bene_add4', bold_one)
        sheet.write(0, 15, 'Add_Details_1', bold_one)
        sheet.write(0, 16, 'Add_Details_2', bold_one)
        sheet.write(0, 17, 'Add_Details_3', bold_one)
        sheet.write(0, 18, 'Add_Details_4', bold_one)
        sheet.write(0, 19, 'Add_Details_5', bold_one)
        sheet.write(0, 20, 'Remarks', bold_one)

        row = 1
        col = 0

        total_tax = 0.0
        pay_method = ""

        for line in batch_id.payment_ids:
            if not line.partner_bank_id.bank_id.bic:
                raise ValidationError(f'Bank/IFSC not set in partner {line.partner_id.name}')
            if not line.batch_payment_id.journal_id.bank_id.bic:
                raise ValidationError(f'Bank/IFSC not set in Journal {line.partner_id.name}')
            if line.batch_payment_id.journal_id.bank_id.bic[0:4] == line.partner_bank_id.bank_id.bic[0:4]:
                pay_method = "I"
            elif line.amount > 200000:
                # current_time = datetime.datetime.now(pytz.timezone('Asia/Kolkata')).time()
                current_time = fields.Datetime.context_timestamp(self, fields.Datetime.now()).time()
                if current_time >= time(15, 0):
                # if current_time >= datetime.time(15, 0):
                    pay_method = "N"
                else:
                    pay_method = "R"
            else:
                pay_method = "N"

            partner_name = re.sub(r"[-\d]+", "", line.partner_id.name)

            sheet.write(row, col, line.batch_payment_id.journal_id.bank_account_id.acc_number)
            sheet.write(row, col + 1, line.partner_bank_id.acc_number or ' ')
            sheet.write(row, col + 2, partner_name or ' ')
            sheet.write(row, col + 3, int(line.amount) or ' ')
            sheet.write(row, col + 4, str(pay_method))
            sheet.write(row, col + 5, line.date.strftime("%d-%b-%Y").upper())
            sheet.write(row, col + 6, line.partner_bank_id.bank_id.bic or ' ')
            sheet.write(row, col + 7, ' ')
            sheet.write(row, col + 8, ' ')
            sheet.write(row, col + 9, ' ')
            sheet.write(row, col + 10, ' ')
            sheet.write(row, col + 11, ' ')
            sheet.write(row, col + 12, ' ')
            sheet.write(row, col + 13, ' ')
            sheet.write(row, col + 14, ' ')
            sheet.write(row, col + 15, ' ')
            sheet.write(row, col + 16, ' ')
            sheet.write(row, col + 17, ' ')
            sheet.write(row, col + 18, ' ')
            sheet.write(row, col + 19, ' ')
            sheet.write(row, col + 20, line.narration)

            row += 1

        # Save the workbook to a file or return it as a response to a web request.
        fp = BytesIO()
        workbook.save(fp)
        self.write(
            {'file_name': base64.b64encode(fp.getvalue()), 'summary_data': file_name})
        fp.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'custom.excel.class',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    # Report for bank format in xls
    def generate_xlsx_report_icici(self):

        batch_id = self.env['account.batch.payment'].browse(self.env.context['active_id'])
        file_name = f"{batch_id.name}_{fields.Date.today().strftime('%d%B%Y')}.xls"
        # file_name = (batch_id.name) + '_' + (batch_id.today_date.strftime('%d%B%Y')) + '.xls'
        bold_one = xlwt.easyxf('font: bold 1, height 240; align: vertical center;')
        title_date_format = xlwt.easyxf(num_format_str='d-MMM-yyyy')

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('TransactionPayeeDetails')
        for col in range(21):
            sheet.col(col).width = 25 * 256  # Setting column width in units of 1/256th of a character width.

        # Header
        row = 0
        col = 0
        # # For Detailed Tds Report and Summary

        pay_method = ""
        for line in batch_id.payment_ids:
            if not line.partner_bank_id.bank_id.bic:
                raise ValidationError(f'Bank/IFSC not set in partner {line.partner_id.name}')
            if not line.batch_payment_id.journal_id.bank_id.bic:
                raise ValidationError(f'Bank/IFSC not set in Journal {line.partner_id.name}')
            if line.batch_payment_id.journal_id.bank_id.bic[0:4] == line.partner_bank_id.bank_id.bic[0:4]:
                pay_method = "I"
            elif line.amount > 200000:
                # current_time = datetime.datetime.now(pytz.timezone('Asia/Kolkata')).time()
                current_time = fields.Datetime.context_timestamp(self, fields.Datetime.now()).time()
                # if current_time >= datetime.time(15, 0):
                if current_time >= time(15, 0):
                    pay_method = "N"
                else:
                    pay_method = "R"
            else:
                pay_method = "N"

            name = re.sub(r"[-\d]+", "", line.partner_id.name).rstrip()
            sheet.write(row, col, str(pay_method))
            sheet.write(row, col + 1, int(line.amount) or ' ')
            sheet.write(row, col + 2, fields.Date.today(), title_date_format)
            sheet.write(row, col + 3, name)
            sheet.write(row, col + 4, line.partner_bank_id.acc_number or ' ')
            sheet.write(row, col + 5, '')
            sheet.write(row, col + 6, line.narration)
            sheet.write(row, col + 7, line.batch_payment_id.journal_id.bank_account_id.acc_number)
            sheet.write(row, col + 8, '')
            sheet.write(row, col + 9, line.partner_bank_id.bank_id.bic)
            sheet.write(row, col + 10, '11')

            row += 1
        # Save the workbook to a file or return it as a response to a web request.
        fp = BytesIO()
        workbook.save(fp)
        self.write(
            {'file_name': base64.b64encode(fp.getvalue()), 'summary_data': file_name})
        fp.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'custom.excel.class',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'target': 'new',
        }


class ResPartnerBankInherit(models.Model):
    _inherit = 'res.partner.bank'

    @api.model
    def create(self, vals):
        res = super(ResPartnerBankInherit, self).create(vals)
        account_number = res.acc_number
        if account_number:
            existing_account = self.search([('acc_number', '=', account_number),('id','!=',res.id)])
            if existing_account:
                voucher_account = self.env['voucher.account.master'].sudo().search([('name', '=', account_number)])
                if voucher_account:
                    return super(ResPartnerBankInherit, self).create(vals)
            if existing_account:
                raise ValidationError('An account with the same account number already exists!!!')
        return res

    @api.onchange('acc_number')
    def validate_account_number(self):
        if self.acc_number:
            existing_account = self.search([('acc_number', '=', self.acc_number)])
            if existing_account:
                voucher_account = self.env['voucher.account.master'].sudo().search([('name', '=', self.acc_number)])
                if not voucher_account:
                    if not self.partner_id.parent_id:
                        raise ValidationError('An account with the same account number already exists!!!')
                    elif self.partner_id.parent_id != existing_account.parent_id or self.partner_id.parent_id != existing_account.partner_id:
                        raise ValidationError('An account with the same account number already exists!!!')





class ResPartnerInherit(models.Model):
    _inherit = "res.partner"

    pan_no = fields.Char('PAN No')
    pan_file_name = fields.Char()
    pan_card = fields.Binary('PAN Card', attachment=True, copy=False)
    cancelled_cheque = fields.Binary('Cancelled Cheque', attachment=True, copy=False)
    check_file_name = fields.Char()
    gst_doc_name = fields.Char()
    gst_doc = fields.Binary('GST', attachment=True, copy=False)
    vat = fields.Char(string='GSTIN', index=True, tracking=True,
                      help="The Tax Identification Number. Values here will be validated based on the country format. You can use '/' to indicate that the partner is not subject to tax.")
    code = fields.Char('Code')

    restrict_partner_ledger = fields.Boolean(string="Restrict Partner Ledger Visibility")


class PurchaseInherit(models.Model):
    _inherit = 'purchase.order'

    def _compute_total_claimable_amt_words(self, amount):
        for rec in self:
            return num2words(str(amount), lang='en_IN').replace('and', '').replace('point', 'and').replace(
                'thous', 'thousand')


class AnalyticAccountInherit(models.Model):
    _inherit = 'account.analytic.account'

    payments_transfer = fields.Float('Payments Transfer')

    # @api.model
    # def create(self, vals):
    #     analytic_name = vals.get('name')
    #     if analytic_name:
    #         existing_account = self.search([('name', '=', analytic_name)])
    #         if existing_account:
    #             raise ValidationError(f'An analytic account {analytic_name} with the same name already exists!!!')
    #     return super(AnalyticAccountInherit, self).create(vals)


class InheritAccountAnalytic(models.Model):
    _inherit = 'account.analytic.account'

    # @api.depends('line_ids.amount')
    # def _compute_debit_credit_balance(self):
    #     Curr = self.env['res.currency']
    #     analytic_line_obj = self.env['account.analytic.line']
    #     domain = [
    #         ('account_id', 'in', self.ids),
    #         ('company_id', 'in', [False] + self.env.companies.ids)
    #     ]
    #     if self._context.get('from_date', False):
    #         domain.append(('date', '>=', self._context['from_date']))
    #     if self._context.get('to_date', False):
    #         domain.append(('date', '<=', self._context['to_date']))
    #
    #     user_currency = self.env.company.currency_id
    #     credit_groups = analytic_line_obj.read_group(
    #         domain=domain + [('amount', '>=', 0.0)],
    #         fields=['account_id', 'currency_id', 'amount'],
    #         groupby=['account_id', 'currency_id'],
    #         lazy=False,
    #     )
    #     data_credit = defaultdict(float)
    #     for l in credit_groups:
    #         data_credit[l['account_id'][0]] += Curr.browse(l['currency_id'][0])._convert(
    #             l['amount'], user_currency, self.env.company, fields.Date.today())
    #
    #     debit_groups = analytic_line_obj.read_group(
    #         domain=domain + [('amount', '<', 0.0)],
    #         fields=['account_id', 'currency_id', 'amount'],
    #         groupby=['account_id', 'currency_id'],
    #         lazy=False,
    #     )
    #     data_debit = defaultdict(float)
    #     for l in debit_groups:
    #         data_debit[l['account_id'][0]] += Curr.browse(l['currency_id'][0])._convert(
    #             l['amount'], user_currency, self.env.company, fields.Date.today())
    #     # bank_journal_ids = self.env['account.journal'].search([('type', '=', 'bank')])
    #     # outgoing_acc_ids = bank_journal_ids.outbound_payment_method_line_ids.mapped('payment_account_id')
    #     # payment_outstanding_account = self.env.company.account_journal_payment_credit_account_id
    #     # outgoing_account_ids = outgoing_acc_ids + payment_outstanding_account
    #     for account in self:
    #         # account_move = self.enc['account.move.line'].search([('account_id.account_type', 'in', ['income', 'expense'])])
    #         debit = 0
    #         credit = 0
    #         pay_t = 0
    #         for rec in account.line_ids:
    #             if rec.move_line_id.account_id.account_type in ['income', 'expense']:
    #                 debit += rec.move_line_id.debit
    #                 credit += rec.move_line_id.credit
    #             if rec.move_line_id.account_id.id in payment_outstanding_account.ids:
    #                 pay_t += rec.move_line_id.credit
    #                 pay_t -= rec.move_line_id.debit
    #         account.debit = debit
    #         account.credit = credit
    #         account.payments_transfer = pay_t
    #         account.balance = account.credit - account.debit